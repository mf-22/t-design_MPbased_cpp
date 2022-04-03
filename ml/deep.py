from util import data_loader, label_generator

import os
import sys
import datetime
import time

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import tensorflow as tf
from sklearn.preprocessing import StandardScaler
import openpyxl


def set_seed(my_seed):
    """ Function which sets seed
        Arg:
            my_seed := Integer that will be the seed of pseudo-random number of type int
                       (greater than or equal to 0 and less than or equal to 2**31)
    """
    #print("seed :", my_seed)
    np.random.seed(my_seed)
    tf.random.set_seed(my_seed)


def deep_learning(repeat=False, shuffle=True):
    """ Functions runnning Deep Learning. The flow is as follows:
            1. Data loading
            2. Set hyper parameters, etc.
            3. Training, and saving the training log
            4. Evaluate test data and save evaluation results
        Args:
            repeat  := bool type flag to display graphs, etc. (default=False)
            shuffle := bool type flag to split train and valid data randomly(default=True)
    """
    ##Get path delimiter ("/" or "\") depending on OS
    SEP = os.sep
    ##Get the current time as a string (e.g.) July 28, 2021 18:40:39 => 20210728184039
    dt_now = datetime.datetime.now()
    dt_index = dt_now.strftime("%Y%m%d%H%M%S")
    ##set the seed
    seed = np.random.randint(2**31)
    set_seed(seed)
    
    ##input the filename
    print("input dataset : ", end=(""))
    data_name = input()
    if not os.path.isdir("datasets" + SEP + data_name):
        print('ERROR: Cannnot find the dataset "{}"'.format(data_name))
        return -1
    ##make a directory path to save deep learning results
    ##(e.g. datasets/dataset1/NN/20210729123234/)
    dir_path = "datasets" + SEP + data_name + SEP + "NN" + SEP + dt_index + SEP
    ##Create a directory to store the results
    os.makedirs(dir_path)

    ##Read the dataset
    train_data, train_label, valid_data, valid_label, test_dataset, test_labelset \
        = data_loader.load_data("NN", data_name, dt_index)
    
    if shuffle:
        ##Combine training and validation data and divide randomly
        num_train_data = train_data.shape[0]
        temp_data = np.concatenate([train_data, valid_data], axis=0)
        temp_label = np.concatenate([train_label, valid_label], axis=0)
        from sklearn.model_selection import train_test_split
        train_data, valid_data, train_label, valid_label = train_test_split(temp_data, temp_label, train_size=num_train_data, random_state=seed)

    ##Scale to the mean become 0, the variance become 1 for each feature in the teacher data
    scaler = StandardScaler()
    ##Calculate shift and scaling ratios using train data and transform
    train_data = scaler.fit_transform(train_data)
    ##validation and test data are transformed with calculated shift and scaling ratios from training dataset
    valid_data = scaler.transform(valid_data)
    test_dataset = [scaler.transform(test_data) for test_data in test_dataset]
    
    print("\n===== Start Deep Learning (Train and validation Step) =====")
    ##Get the dimension of the feature vector
    input_node_num = train_data.shape[1]
    ##Set hyperparameters of deep learning
    hidden_node_num = input_node_num #number of hidden node
    hidden_layer_num = 1 #number of hidden layer
    epoch = 100
    batch = 128

    ##Generate empty NN model
    model = tf.keras.Sequential()
    ##add input layer
    model.add(tf.keras.layers.Dense(input_node_num, input_dim=input_node_num))
    #model.add(tf.keras.layers.BatchNormalization())
    model.add(tf.keras.layers.Activation("relu"))
    ##add hidden layer
    for _ in range(hidden_layer_num):
        #model.add(tf.keras.layers.Dense(hidden_node_num, input_dim=input_node_num))
        model.add(tf.keras.layers.Dense(hidden_node_num))
        #model.add(tf.keras.layers.BatchNormalization())
        model.add(tf.keras.layers.Activation("relu"))
        #model.add(tf.keras.layers.Dropout(0.5))
    ##add output layer
    #model.add(tf.keras.layers.Dense(1, input_dim=input_node_num, activation="sigmoid"))
    model.add(tf.keras.layers.Dense(1, activation="sigmoid"))

    ##set learing rate(keras's adam default = 0.001)
    learning_rate = 1.0 * 10**(-5)
    ##set the optimizer, loss function and others
    opt = tf.keras.optimizers.Adam(lr=learning_rate)
    model.compile(optimizer=opt,
                  loss="binary_crossentropy",
                  metrics=["accuracy"])
    
    ##output the NN model we created
    print(model.summary())
    ##Draw and save the created model to a text file
    sys.stdout = open(dir_path+"model.txt", mode="w")
    print(model.summary())
    sys.stdout = sys.__stdout__
    ##Save as png as well
    tf.keras.utils.plot_model(model, to_file=dir_path+"model.png")

    ##callbacks
    ##Callback to save the model at each epoch. However, when val_accuracy is better.
    mc_cb = tf.keras.callbacks.ModelCheckpoint(
        filepath=dir_path+"best_model",
        save_weights_only=False,
        monitor="val_accuracy",
        verbose=1,
        save_best_only=True,
    )
    ##Callback to terminate learning when the value being monitored does not improve between specific epochs
    """
    es_cb = tf.keras.callbacks.EarlyStopping(
        monitor="val_accuracy",
        mode="max",
        min_delta=0.001,
        patience=epoch//4,
        verbose=1,
        restore_best_weights=True #Load weights at the end of training when the value being monitored is the best
    )"""

    ##Train the model with training data and obtain the training process
    start = time.perf_counter()
    history = model.fit(
        train_data,
        train_label,
        epochs=epoch,
        batch_size=batch,
        validation_data=(valid_data, valid_label),
        callbacks=[mc_cb]
        #callbacks=[mc_cb, es_cb]
    )
    finish = time.perf_counter()
    ##Output time spent on training
    print("fit time : {}[s]".format(finish-start))

    ##Save the process for each epoch in the learning to an excel file at once
    hist_df = pd.DataFrame(history.history)
    hist_df.to_excel(dir_path+"results.xlsx", sheet_name="history")

    ##Draw a graph of LOSS and ACCURACY at each epoch in the study.
    ##Get the value of the error function for each epoch
    loss = history.history["loss"]
    val_loss = history.history["val_loss"]
    ##Create an object with elements from 1 to epoch
    epochs = range(1, len(loss)+1)
    ##plot
    plt.plot(epochs, loss, "bo", label="Training loss") #Plot the values of the error function for the training data
    plt.plot(epochs, val_loss, "b", label="Validation loss") #Plot the values of the error function for the validation data
    plt.title("Training and validation loss") #graph title
    plt.xlabel("Epochs") #label of horizontal axis
    plt.ylabel("Loss") #label of vertical axis
    plt.legend() #output legend
    plt.savefig(dir_path+"loss.png") #save as png
    if not repeat:
        ## When deep learning is repeated 10 times or so, it is not displayed.
        plt.show() #Displaying graphs in a window
    
    ## 各エポックごとの識別精度の値を取得
    acc = history.history["accuracy"]
    val_acc = history.history["val_accuracy"]
    plt.figure() #clean the graph object
    plt.plot(epochs, acc, "bo", label="Training acc") #Plot the values of the error function for the training data
    plt.plot(epochs, val_acc, "b", label="Validation acc") #Plot the values of the error function for the validation data
    plt.title("Training and validation accuracy") #graph title
    plt.xlabel("Epochs") #label of horizontal axis
    plt.ylabel("Accuracy") #label of vertical axis
    plt.legend() #output legend
    plt.savefig(dir_path+"acc.png") #save as png
    if not repeat:
        plt.show() #Displaying graphs in a window
    
    ##Save parameters of deep learning
    with open(dir_path+"train_parameters.txt", mode="w") as f:
        f.write("shuffle        : {}\n".format(shuffle))
        f.write("seed           : {}\n".format(seed))
        f.write("optimizer      : {}\n".format(opt))
        f.write("learning rate  : {}\n".format(learning_rate))
        f.write("#hidden layers : {}\n".format(hidden_layer_num))
        f.write("input_dim      : {}\n".format(input_node_num))
        f.write("#hidden node   : {}\n".format(hidden_node_num))
        f.write("epoch          : {}\n".format(epoch))
        f.write("batch size     : {}\n".format(batch))
        f.write("fit time       : {}\n".format(finish-start))

    ##Make a folder to save how training and validation data are identified
    os.makedirs(dir_path+"predicted_labels")
    ##Input training data to trained NN, convert to labels, and save.
    train_predicted_label = label_generator.scaler_to_label(model.predict(train_data))
    np.save(dir_path+"predicted_labels/train_predicted_label.npy", train_predicted_label)
    ##Input validation data into trained NN, convert to labels, and save
    valid_predicted_label = label_generator.scaler_to_label(model.predict(valid_data))
    np.save(dir_path+"predicted_labels/valid_predicted_label.npy", valid_predicted_label)
    
    ##Evaluate test data
    print("\n\n===== Start test using NN (Test step) =====")
    ##Test step is done on the model when it was the best (val_acc was high) in the training, load for that.
    ##EarlyStopping callback does the same thing.
    model = tf.keras.models.load_model(dir_path+"best_model")

    ##Save the test results, etc. in the Excel file where you saved the training log earlier.
    wb = openpyxl.load_workbook(dir_path + "results.xlsx")
    ##Write in a new sheet "test_results"
    test_ws = wb.create_sheet(title="test_results")
    test_ws["B1"] = "accuracy"
    test_ws["C1"] = "count0"
    test_ws["D1"] = "count1"

    for i in range(len(test_dataset)):
        ##Extracting test data from the list of test dataset
        test_data = test_dataset[i]
        test_label = test_labelset[i]
        print("\n** test{} **".format(i+1))

        ##get identification accuracy of test data
        data_loss, data_acc = model.evaluate(test_data, test_label)
        
        ##Input test data into trained NN, convert to labels, and save.
        test_predicted_label = label_generator.scaler_to_label(model.predict(test_data))
        np.save(dir_path+"predicted_labels/test{}_predicted_label.npy".format(i+1), test_predicted_label)
        ##Get the number of NN predicted results of 0 and 1, respectively
        count_0 = np.count_nonzero(test_predicted_label==0)
        count_1 = np.count_nonzero(test_predicted_label==1)
        print("acc : {}, label0 : {}, label1 : {}".format(data_acc, count_0, count_1))

        ##Write the results on the excel sheet
        test_ws.cell(row=i+2, column=1, value="test{}".format(i+1))
        test_ws.cell(row=i+2, column=2, value=data_acc)
        test_ws.cell(row=i+2, column=3, value=count_0)
        test_ws.cell(row=i+2, column=4, value=count_1)
    
    ## Save the writing to excel sheet and exit
    wb.save(dir_path+"results.xlsx")


if __name__ == "__main__":
    deep_learning(repeat=True, shuffle=True)